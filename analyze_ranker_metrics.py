import argparse
import json
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from statistics import mean, median
from typing import Iterable

import matplotlib.pyplot as plt
from openpyxl import load_workbook


DEFAULT_KS = [1, 3, 5, 10, 20, 50, 100]
DEFAULT_POSITIVE_LABELS = {"TopAnswer"}
DEFAULT_NEGATIVE_LABELS = {"NeverShow"}
RANKERS = ("ARAGORN_RANKER", "ARAX_RANKER")
RANKER_COLUMNS = {
    "ARAGORN_RANKER": "aragorn_rank",
    "ARAX_RANKER": "arax_rank",
}
RANKER_COLORS = {
    "ARAGORN_RANKER": "#1f77b4",
    "ARAX_RANKER": "#d62728",
}


@dataclass
class EdgeRecord:
    qid: int
    ara: str
    edge_id: str | None
    subject_id: str | None
    subject_name: str | None
    object_id: str | None
    object_name: str | None
    predicate: str | None
    labels: set[str]
    ranks: dict[str, int | None]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Analyze ARAGORN_RANKER and ARAX_RANKER performance from a comparison "
            "workbook. Each workbook row is treated as one ranked edge."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("results/ranker_comparison_test_queries.xlsx"),
        help="Path to the comparison workbook.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Results sheet name. If omitted, the script auto-detects it.",
    )
    parser.add_argument(
        "--ks",
        type=int,
        nargs="+",
        default=DEFAULT_KS,
        help="Top-k cutoffs for hits@k and related metrics.",
    )
    parser.add_argument(
        "--positive-labels",
        nargs="+",
        default=sorted(DEFAULT_POSITIVE_LABELS),
        help="Labels treated as desirable edges.",
    )
    parser.add_argument(
        "--negative-labels",
        nargs="+",
        default=sorted(DEFAULT_NEGATIVE_LABELS),
        help="Labels treated as undesirable edges.",
    )
    parser.add_argument(
        "--json-output",
        type=Path,
        default=None,
        help="Optional path for machine-readable JSON output.",
    )
    parser.add_argument(
        "--duplicate-report",
        type=Path,
        default=None,
        help=(
            "Optional path for a JSON report of duplicate edge_ids within the same qid. "
            "If omitted, a default path next to the workbook is used."
        ),
    )
    parser.add_argument(
        "--plot-dir",
        type=Path,
        default=None,
        help="Directory where PNG comparison plots should be written.",
    )
    parser.add_argument(
        "--show-group-examples",
        type=int,
        default=5,
        help="Number of strongest per-group differences to print.",
    )
    return parser.parse_args()


def normalize_header(value: object) -> str:
    return str(value).strip().lower().replace(" ", "_")


def find_results_sheet(workbook, requested_name: str | None):
    if requested_name:
        if requested_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{requested_name}' not found in workbook")
        return workbook[requested_name]

    normalized = {normalize_header(name): name for name in workbook.sheetnames}
    for candidate in ("ara_ranker_results", "ara_ranker_result", "ara_ranker_results_"):
        if candidate in normalized:
            return workbook[normalized[candidate]]

    for name in workbook.sheetnames:
        normalized_name = normalize_header(name)
        if "ranker" in normalized_name and "result" in normalized_name:
            return workbook[name]

    raise ValueError(
        f"Could not auto-detect results sheet. Workbook sheets: {workbook.sheetnames}"
    )


def required_column_map(headers: Iterable[object]) -> dict[str, int]:
    normalized_to_index = {
        normalize_header(header): index for index, header in enumerate(headers)
    }
    required = {
        "qid": "qid",
        "ara": "ara",
        "expected_output": "expected_output",
        "subject_id": "subject_id",
        "subject_name": "subject_name",
        "object_id": "object_id",
        "object_name": "object_name",
        "predicate": "predicate",
        "edge_id": "edge_id",
        "aragorn_rank": "aragorn_rank",
        "arax_rank": "arax_rank",
    }
    missing = [name for name in required if name not in normalized_to_index]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    return {
        canonical: normalized_to_index[normalized]
        for canonical, normalized in required.items()
    }


def maybe_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def load_edge_records(
    workbook_path: Path,
    sheet_name: str | None,
) -> tuple[list[EdgeRecord], dict[str, object], list[dict[str, object]]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = find_results_sheet(workbook, sheet_name)
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    columns = required_column_map(headers)

    records: list[EdgeRecord] = []
    query_edge_keys = set()
    query_edge_occurrences: dict[tuple[int, object], list[dict[str, object]]] = (
        defaultdict(list)
    )

    for row_number, row in enumerate(rows, start=2):
        qid = row[columns["qid"]]
        if qid is None:
            continue
        qid = int(qid)
        edge_id = row[columns["edge_id"]]
        query_edge_key = (qid, edge_id)
        query_edge_keys.add(query_edge_key)
        query_edge_occurrences[query_edge_key].append(
            {
                "row_number": row_number,
                "ARA": str(row[columns["ara"]]),
                "predicate": row[columns["predicate"]],
                "subject_id": row[columns["subject_id"]],
                "subject_name": row[columns["subject_name"]],
                "object_id": row[columns["object_id"]],
                "object_name": row[columns["object_name"]],
                "expected_output": row[columns["expected_output"]],
                "aragorn_rank": maybe_int(row[columns["aragorn_rank"]]),
                "arax_rank": maybe_int(row[columns["arax_rank"]]),
            }
        )

        label = row[columns["expected_output"]]
        records.append(
            EdgeRecord(
                qid=qid,
                ara=str(row[columns["ara"]]),
                edge_id=edge_id,
                subject_id=row[columns["subject_id"]],
                subject_name=row[columns["subject_name"]],
                object_id=row[columns["object_id"]],
                object_name=row[columns["object_name"]],
                predicate=row[columns["predicate"]],
                labels={str(label)} if label else set(),
                ranks={
                    "ARAGORN_RANKER": maybe_int(row[columns["aragorn_rank"]]),
                    "ARAX_RANKER": maybe_int(row[columns["arax_rank"]]),
                },
            )
        )

    duplicate_report = []
    duplicate_row_count = 0
    for (qid, edge_id), rows_for_key in sorted(query_edge_occurrences.items()):
        if len(rows_for_key) <= 1:
            continue
        duplicate_row_count += len(rows_for_key)
        duplicate_report.append(
            {
                "qid": qid,
                "edge_id": edge_id,
                "occurrences": len(rows_for_key),
                "aras": sorted({row["ARA"] for row in rows_for_key}),
                "rows": rows_for_key,
            }
        )

    metadata = {
        "sheet": sheet.title,
        "raw_rows": len(records),
        "edge_records": len(records),
        "distinct_query_edge_keys": len(query_edge_keys),
        "duplicate_query_edge_ids": len(duplicate_report),
        "duplicate_query_edge_rows": duplicate_row_count,
    }
    return records, metadata, duplicate_report


def label_category(
    labels: set[str],
    positive_labels: set[str],
    negative_labels: set[str],
) -> str:
    has_positive = bool(labels & positive_labels)
    has_negative = bool(labels & negative_labels)
    if has_positive and has_negative:
        return "ambiguous"
    if has_positive:
        return "positive"
    if has_negative:
        return "negative"
    return "other"


def ranked_records(records: list[EdgeRecord], ranker_name: str) -> list[EdgeRecord]:
    return sorted(
        (record for record in records if record.ranks[ranker_name] is not None),
        key=lambda record: (
            record.ranks[ranker_name],
            record.edge_id or "",
            record.subject_id or "",
            record.object_id or "",
        ),
    )


def summarize_group_metrics(
    group_records: list[EdgeRecord],
    ranker_name: str,
    ks: list[int],
    positive_labels: set[str],
    negative_labels: set[str],
) -> dict[int, dict[str, float | int | None]]:
    categories = {
        id(record): label_category(record.labels, positive_labels, negative_labels)
        for record in group_records
    }

    total_positive = sum(
        1 for record in group_records if categories[id(record)] == "positive"
    )
    total_negative = sum(
        1 for record in group_records if categories[id(record)] == "negative"
    )
    total_ambiguous = sum(
        1 for record in group_records if categories[id(record)] == "ambiguous"
    )

    summary = {}
    for k in ks:
        in_top_k = [
            record
            for record in group_records
            if record.ranks[ranker_name] is not None and record.ranks[ranker_name] <= k
        ]
        positive_hits = sum(
            1 for record in in_top_k if categories[id(record)] == "positive"
        )
        negative_hits = sum(
            1 for record in in_top_k if categories[id(record)] == "negative"
        )
        false_negatives = total_positive - positive_hits
        precision = (
            positive_hits / (positive_hits + negative_hits)
            if (positive_hits + negative_hits)
            else None
        )
        positive_recall = positive_hits / total_positive if total_positive else None
        f1 = (
            2 * precision * positive_recall / (precision + positive_recall)
            if precision is not None
            and positive_recall is not None
            and (precision + positive_recall)
            else None
        )
        summary[k] = {
            "top_k_size": len(in_top_k),
            "positive_hits": positive_hits,
            "negative_hits": negative_hits,
            "false_negatives": false_negatives,
            "total_positive": total_positive,
            "total_negative": total_negative,
            "total_ambiguous": total_ambiguous,
            "positive_recall": positive_recall,
            "precision": precision,
            "f1": f1,
            "negative_specificity": (
                (total_negative - negative_hits) / total_negative
                if total_negative
                else None
            ),
            "negative_exposure": (
                negative_hits / total_negative if total_negative else None
            ),
        }
    return summary


def empty_aggregate(ks: list[int]) -> dict[int, dict[str, object]]:
    return {
        k: {
            "positive_hits": 0,
            "negative_hits": 0,
            "false_negatives": 0,
            "top_k_size": 0,
            "positive_totals": 0,
            "negative_totals": 0,
            "precisions": [],
            "positive_recalls": [],
            "f1_scores": [],
            "negative_specificities": [],
            "negative_exposures": [],
            "eligible_positive_groups": 0,
            "eligible_precision_groups": 0,
            "eligible_f1_groups": 0,
            "eligible_negative_groups": 0,
        }
        for k in ks
    }


def finalize_aggregate_map(
    aggregate: dict[int, dict[str, object]],
) -> dict[int, dict[str, object]]:
    finalized = {}
    for k, metrics in aggregate.items():
        tp = metrics["positive_hits"]
        fp = metrics["negative_hits"]
        fn = metrics["false_negatives"]
        positive_totals = metrics["positive_totals"]
        negative_totals = metrics["negative_totals"]
        precision_micro = tp / (tp + fp) if (tp + fp) else None
        recall_micro = tp / positive_totals if positive_totals else None
        f1_micro = 2 * tp / (2 * tp + fp + fn) if (2 * tp + fp + fn) else None
        finalized[k] = {
            "positive_hits": tp,
            "negative_hits": fp,
            "false_negatives": fn,
            "top_k_size": metrics["top_k_size"],
            "positive_totals": positive_totals,
            "negative_totals": negative_totals,
            "precision_micro": precision_micro,
            "positive_recall_micro": recall_micro,
            "f1_micro": f1_micro,
            "negative_specificity_micro": (
                (negative_totals - fp) / negative_totals if negative_totals else None
            ),
            "negative_exposure_micro": (
                fp / negative_totals if negative_totals else None
            ),
            "positive_hits_per_group_mean": (
                tp / metrics["eligible_positive_groups"]
                if metrics["eligible_positive_groups"]
                else None
            ),
            "negative_hits_per_group_mean": (
                fp / metrics["eligible_negative_groups"]
                if metrics["eligible_negative_groups"]
                else None
            ),
            "precision_macro": (
                mean(metrics["precisions"]) if metrics["precisions"] else None
            ),
            "positive_recall_macro": (
                mean(metrics["positive_recalls"])
                if metrics["positive_recalls"]
                else None
            ),
            "f1_macro": mean(metrics["f1_scores"]) if metrics["f1_scores"] else None,
            "negative_specificity_macro": (
                mean(metrics["negative_specificities"])
                if metrics["negative_specificities"]
                else None
            ),
            "negative_exposure_macro": (
                mean(metrics["negative_exposures"])
                if metrics["negative_exposures"]
                else None
            ),
            "eligible_positive_groups": metrics["eligible_positive_groups"],
            "eligible_precision_groups": metrics["eligible_precision_groups"],
            "eligible_f1_groups": metrics["eligible_f1_groups"],
            "eligible_negative_groups": metrics["eligible_negative_groups"],
        }
    return finalized


def summarize_pairwise_margin_buckets(
    pairwise_margin_buckets: dict[int, dict[str, dict[str, list[int]]]],
) -> dict[int, dict[str, dict[str, dict[str, float | int | None]]]]:
    pairwise_margin_summary = {}
    for k, metric_buckets in pairwise_margin_buckets.items():
        pairwise_margin_summary[k] = {}
        for metric_name, ranker_buckets in metric_buckets.items():
            pairwise_margin_summary[k][metric_name] = {}
            for ranker, values in ranker_buckets.items():
                pairwise_margin_summary[k][metric_name][ranker] = {
                    "win_count": len(values),
                    "mean_margin": mean(values) if values else None,
                    "median_margin": median(values) if values else None,
                    "total_margin": sum(values),
                }
    return pairwise_margin_summary


def aggregate_metric_table(
    records: list[EdgeRecord],
    ks: list[int],
    positive_labels: set[str],
    negative_labels: set[str],
) -> tuple[
    dict[str, dict[int, dict[str, object]]],
    dict[int, dict[str, dict[str, int]]],
    dict[int, dict[str, dict[str, dict[str, float | int | None]]]],
    dict[str, dict[str, dict[int, dict[str, object]]]],
    dict[str, dict[int, dict[str, dict[str, int]]]],
    dict[str, dict[int, dict[str, dict[str, dict[str, float | int | None]]]]],
    dict[str, list[dict[str, object]]],
    dict[str, dict[str, list[dict[str, object]]]],
]:
    grouped: dict[tuple[str, int], list[EdgeRecord]] = defaultdict(list)
    for record in records:
        grouped[(record.ara, record.qid)].append(record)

    overall = {ranker: empty_aggregate(ks) for ranker in RANKERS}
    by_ara = defaultdict(lambda: {ranker: empty_aggregate(ks) for ranker in RANKERS})
    pairwise_overall = {
        k: {
            "positive_hits": {
                "ARAGORN_RANKER": 0,
                "ARAX_RANKER": 0,
                "tie": 0,
                "tie_zero": 0,
                "tie_nonzero": 0,
            },
            "negative_hits": {
                "ARAGORN_RANKER": 0,
                "ARAX_RANKER": 0,
                "tie": 0,
                "tie_zero": 0,
                "tie_nonzero": 0,
            },
        }
        for k in ks
    }
    pairwise_by_ara = defaultdict(
        lambda: {
            k: {
                "positive_hits": {
                    "ARAGORN_RANKER": 0,
                    "ARAX_RANKER": 0,
                    "tie": 0,
                    "tie_zero": 0,
                    "tie_nonzero": 0,
                },
                "negative_hits": {
                    "ARAGORN_RANKER": 0,
                    "ARAX_RANKER": 0,
                    "tie": 0,
                    "tie_zero": 0,
                    "tie_nonzero": 0,
                },
            }
            for k in ks
        }
    )
    strongest_positive_examples = {ranker: [] for ranker in RANKERS}
    strongest_negative_examples = {
        "fewest_nevershow": {ranker: [] for ranker in RANKERS},
        "most_nevershow": {ranker: [] for ranker in RANKERS},
    }
    pairwise_margin_overall = {
        k: {
            "positive_hits": {ranker: [] for ranker in RANKERS},
            "negative_hits": {ranker: [] for ranker in RANKERS},
        }
        for k in ks
    }
    pairwise_margin_by_ara = defaultdict(
        lambda: {
            k: {
                "positive_hits": {ranker: [] for ranker in RANKERS},
                "negative_hits": {ranker: [] for ranker in RANKERS},
            }
            for k in ks
        }
    )

    for (ara, qid), group_records in grouped.items():
        per_ranker = {
            ranker: summarize_group_metrics(
                group_records,
                ranker,
                ks,
                positive_labels,
                negative_labels,
            )
            for ranker in RANKERS
        }

        for ranker in RANKERS:
            for k, metrics in per_ranker[ranker].items():
                target = overall[ranker][k]
                target["positive_hits"] += metrics["positive_hits"]
                target["negative_hits"] += metrics["negative_hits"]
                target["false_negatives"] += metrics["false_negatives"]
                target["top_k_size"] += metrics["top_k_size"]
                target["positive_totals"] += metrics["total_positive"]
                target["negative_totals"] += metrics["total_negative"]
                if metrics["precision"] is not None:
                    target["precisions"].append(metrics["precision"])
                    target["eligible_precision_groups"] += 1
                if metrics["positive_recall"] is not None:
                    target["positive_recalls"].append(metrics["positive_recall"])
                    target["eligible_positive_groups"] += 1
                if metrics["f1"] is not None:
                    target["f1_scores"].append(metrics["f1"])
                    target["eligible_f1_groups"] += 1
                if metrics["negative_specificity"] is not None:
                    target["negative_specificities"].append(
                        metrics["negative_specificity"]
                    )
                    target["negative_exposures"].append(metrics["negative_exposure"])
                    target["eligible_negative_groups"] += 1

                ara_target = by_ara[ara][ranker][k]
                ara_target["positive_hits"] += metrics["positive_hits"]
                ara_target["negative_hits"] += metrics["negative_hits"]
                ara_target["false_negatives"] += metrics["false_negatives"]
                ara_target["top_k_size"] += metrics["top_k_size"]
                ara_target["positive_totals"] += metrics["total_positive"]
                ara_target["negative_totals"] += metrics["total_negative"]
                if metrics["precision"] is not None:
                    ara_target["precisions"].append(metrics["precision"])
                    ara_target["eligible_precision_groups"] += 1
                if metrics["positive_recall"] is not None:
                    ara_target["positive_recalls"].append(metrics["positive_recall"])
                    ara_target["eligible_positive_groups"] += 1
                if metrics["f1"] is not None:
                    ara_target["f1_scores"].append(metrics["f1"])
                    ara_target["eligible_f1_groups"] += 1
                if metrics["negative_specificity"] is not None:
                    ara_target["negative_specificities"].append(
                        metrics["negative_specificity"]
                    )
                    ara_target["negative_exposures"].append(
                        metrics["negative_exposure"]
                    )
                    ara_target["eligible_negative_groups"] += 1

        for k in ks:
            aragorn = per_ranker["ARAGORN_RANKER"][k]
            arax = per_ranker["ARAX_RANKER"][k]

            if aragorn["total_positive"] or arax["total_positive"]:
                if aragorn["positive_hits"] > arax["positive_hits"]:
                    winner = "ARAGORN_RANKER"
                elif aragorn["positive_hits"] < arax["positive_hits"]:
                    winner = "ARAX_RANKER"
                else:
                    winner = "tie"
                pairwise_overall[k]["positive_hits"][winner] += 1
                pairwise_by_ara[ara][k]["positive_hits"][winner] += 1
                if winner == "tie":
                    tie_bucket = (
                        "tie_zero" if aragorn["positive_hits"] == 0 else "tie_nonzero"
                    )
                    pairwise_overall[k]["positive_hits"][tie_bucket] += 1
                    pairwise_by_ara[ara][k]["positive_hits"][tie_bucket] += 1

                delta = aragorn["positive_hits"] - arax["positive_hits"]
                if delta != 0:
                    winner = "ARAGORN_RANKER" if delta > 0 else "ARAX_RANKER"
                    pairwise_margin_overall[k]["positive_hits"][winner].append(
                        abs(delta)
                    )
                    pairwise_margin_by_ara[ara][k]["positive_hits"][winner].append(
                        abs(delta)
                    )
                    strongest_positive_examples[winner].append(
                        {
                            "qid": qid,
                            "ARA": ara,
                            "k": k,
                            "positive_hit_delta": abs(delta),
                            "aragorn_positive_hits": aragorn["positive_hits"],
                            "arax_positive_hits": arax["positive_hits"],
                            "total_positive": max(
                                aragorn["total_positive"], arax["total_positive"]
                            ),
                        }
                    )

            if aragorn["total_negative"] or arax["total_negative"]:
                if aragorn["negative_hits"] < arax["negative_hits"]:
                    winner = "ARAGORN_RANKER"
                elif aragorn["negative_hits"] > arax["negative_hits"]:
                    winner = "ARAX_RANKER"
                else:
                    winner = "tie"
                pairwise_overall[k]["negative_hits"][winner] += 1
                pairwise_by_ara[ara][k]["negative_hits"][winner] += 1
                if winner == "tie":
                    tie_bucket = (
                        "tie_zero" if aragorn["negative_hits"] == 0 else "tie_nonzero"
                    )
                    pairwise_overall[k]["negative_hits"][tie_bucket] += 1
                    pairwise_by_ara[ara][k]["negative_hits"][tie_bucket] += 1

                delta = aragorn["negative_hits"] - arax["negative_hits"]
                if delta != 0:
                    fewer_winner = "ARAGORN_RANKER" if delta < 0 else "ARAX_RANKER"
                    more_winner = "ARAGORN_RANKER" if delta > 0 else "ARAX_RANKER"
                    pairwise_margin_overall[k]["negative_hits"][fewer_winner].append(
                        abs(delta)
                    )
                    pairwise_margin_by_ara[ara][k]["negative_hits"][
                        fewer_winner
                    ].append(abs(delta))
                    example = {
                        "qid": qid,
                        "ARA": ara,
                        "k": k,
                        "negative_hit_delta": abs(delta),
                        "aragorn_negative_hits": aragorn["negative_hits"],
                        "arax_negative_hits": arax["negative_hits"],
                        "total_negative": max(
                            aragorn["total_negative"], arax["total_negative"]
                        ),
                    }
                    strongest_negative_examples["fewest_nevershow"][
                        fewer_winner
                    ].append(example)
                    strongest_negative_examples["most_nevershow"][more_winner].append(
                        example
                    )

    finalized_overall = {
        ranker: finalize_aggregate_map(aggregate)
        for ranker, aggregate in overall.items()
    }
    finalized_by_ara = {
        ara: {
            ranker: finalize_aggregate_map(aggregate)
            for ranker, aggregate in ranker_aggregates.items()
        }
        for ara, ranker_aggregates in by_ara.items()
    }
    sorted_positive_examples = {
        ranker: sorted(
            examples,
            key=lambda item: (item["positive_hit_delta"], -item["k"]),
            reverse=True,
        )
        for ranker, examples in strongest_positive_examples.items()
    }
    sorted_negative_examples = {
        metric_name: {
            ranker: sorted(
                examples,
                key=lambda item: (item["negative_hit_delta"], -item["k"]),
                reverse=True,
            )
            for ranker, examples in ranker_examples.items()
        }
        for metric_name, ranker_examples in strongest_negative_examples.items()
    }
    pairwise_margin_summary = summarize_pairwise_margin_buckets(pairwise_margin_overall)
    pairwise_margin_by_ara_summary = {
        ara: summarize_pairwise_margin_buckets(margins)
        for ara, margins in pairwise_margin_by_ara.items()
    }
    return (
        finalized_overall,
        dict(pairwise_overall),
        pairwise_margin_summary,
        finalized_by_ara,
        dict(pairwise_by_ara),
        pairwise_margin_by_ara_summary,
        sorted_positive_examples,
        sorted_negative_examples,
    )


def format_rate(value: float | None) -> str:
    if value is None:
        return "NA"
    return f"{value:.3f}"


def print_dataset_summary(
    metadata: dict[str, object], records: list[EdgeRecord]
) -> None:
    grouped = {(record.ara, record.qid) for record in records}
    label_counter = Counter()
    ambiguous = 0
    for record in records:
        if not record.labels:
            label_counter["Unlabeled"] += 1
        else:
            for label in sorted(record.labels):
                label_counter[label] += 1
            if len(record.labels) > 1:
                ambiguous += 1

    print("Dataset summary")
    print(f"  Sheet: {metadata['sheet']}")
    print(f"  Raw rows: {metadata['raw_rows']}")
    print(f"  Edge records analyzed: {metadata['edge_records']}")
    print(f"  Distinct (qid, edge_id) keys: {metadata['distinct_query_edge_keys']}")
    print(
        f"  Duplicate edge_ids within the same qid: {metadata['duplicate_query_edge_ids']}"
    )
    print(
        f"  Rows participating in those duplicates: {metadata['duplicate_query_edge_rows']}"
    )
    print(f"  Distinct (ARA, qid) groups: {len(grouped)}")
    print(f"  Records carrying multiple labels: {ambiguous}")
    print("  Label counts:")
    for label, count in sorted(label_counter.items()):
        print(f"    {label}: {count}")


def print_duplicate_summary(duplicate_report: list[dict[str, object]]) -> None:
    print("Duplicate edge_ids within the same qid")
    if not duplicate_report:
        print("  none")
        print()
        return

    for item in duplicate_report:
        print(
            f"  qid={item['qid']} edge_id={item['edge_id']} occurrences={item['occurrences']} "
            f"ARAs={','.join(item['aras'])}"
        )
    print()


def print_metric_block(
    title: str,
    metrics_by_ranker: dict[str, dict[int, dict[str, object]]],
    ks: list[int],
) -> None:
    print(title)
    for ranker, by_k in metrics_by_ranker.items():
        print(f"  {ranker}")
        print(
            "    k  pos_hits  neg_hits  precision_micro  recall_micro  f1_micro  "
            "precision_macro  recall_macro  f1_macro"
        )
        for k in ks:
            row = by_k[k]
            print(
                "    "
                f"{k:<3}"
                f"{row['positive_hits']:<10}"
                f"{row['negative_hits']:<10}"
                f"{format_rate(row['precision_micro']):<17}"
                f"{format_rate(row['positive_recall_micro']):<14}"
                f"{format_rate(row['f1_micro']):<10}"
                f"{format_rate(row['precision_macro']):<17}"
                f"{format_rate(row['positive_recall_macro']):<14}"
                f"{format_rate(row['f1_macro'])}"
            )
        print("    Supplemental negative metrics")
        print("    k  neg_total  neg_specificity_micro  neg_specificity_macro")
        for k in ks:
            row = by_k[k]
            print(
                "    "
                f"{k:<3}"
                f"{row['negative_totals']:<11}"
                f"{format_rate(row['negative_specificity_micro']):<24}"
                f"{format_rate(row['negative_specificity_macro'])}"
            )
        print()


def print_pairwise_block(
    title: str,
    pairwise: dict[int, dict[str, dict[str, int]]],
    ks: list[int],
) -> None:
    print(title)
    print("  Positive hits wins: more TopAnswer edges in top-k is better")
    print("    k  aragorn  arax  tie  tie_zero  tie_nonzero")
    for k in ks:
        row = pairwise[k]["positive_hits"]
        print(
            f"    {k:<3}{row['ARAGORN_RANKER']:<9}{row['ARAX_RANKER']:<6}"
            f"{row['tie']:<5}{row['tie_zero']:<10}{row['tie_nonzero']}"
        )
    print("  Negative hits wins: fewer NeverShow edges in top-k is better")
    print("    k  aragorn  arax  tie  tie_zero  tie_nonzero")
    for k in ks:
        row = pairwise[k]["negative_hits"]
        print(
            f"    {k:<3}{row['ARAGORN_RANKER']:<9}{row['ARAX_RANKER']:<6}"
            f"{row['tie']:<5}{row['tie_zero']:<10}{row['tie_nonzero']}"
        )
    print()


def print_pairwise_margin_block(
    title: str,
    pairwise_margin: dict[int, dict[str, dict[str, dict[str, float | int | None]]]],
    ks: list[int],
) -> None:
    print(title)
    print("  Positive hits margin: average extra TopAnswer hits among groups won")
    print("    k  aragorn_mean  arax_mean  aragorn_wins  arax_wins")
    for k in ks:
        row = pairwise_margin[k]["positive_hits"]
        print(
            f"    {k:<3}"
            f"{format_rate(row['ARAGORN_RANKER']['mean_margin']):<14}"
            f"{format_rate(row['ARAX_RANKER']['mean_margin']):<11}"
            f"{row['ARAGORN_RANKER']['win_count']:<14}"
            f"{row['ARAX_RANKER']['win_count']}"
        )
    print("  Negative hits margin: average fewer NeverShow hits among groups won")
    print("    k  aragorn_mean  arax_mean  aragorn_wins  arax_wins")
    for k in ks:
        row = pairwise_margin[k]["negative_hits"]
        print(
            f"    {k:<3}"
            f"{format_rate(row['ARAGORN_RANKER']['mean_margin']):<14}"
            f"{format_rate(row['ARAX_RANKER']['mean_margin']):<11}"
            f"{row['ARAGORN_RANKER']['win_count']:<14}"
            f"{row['ARAX_RANKER']['win_count']}"
        )
    print()


def print_examples(
    title: str,
    examples_by_ranker: dict[str, list[dict[str, object]]],
    limit: int,
    delta_field: str,
) -> None:
    print(title)
    for ranker, examples in examples_by_ranker.items():
        print(f"  {ranker}")
        if not examples:
            print("    none")
            continue
        for example in examples[:limit]:
            print(
                f"    ARA={example['ARA']} qid={example['qid']} k={example['k']} "
                f"delta={example[delta_field]} details={json.dumps(example, sort_keys=True)}"
            )
    print()


def plot_line_panel(
    axes,
    row_index: int,
    row_title: str,
    metric_by_ranker: dict[str, dict[int, dict[str, object]]],
    ks: list[int],
    left_key: str,
    left_title: str,
    left_ylabel: str,
    right_key: str,
    right_title: str,
    right_ylabel: str,
) -> None:
    left_ax = axes[row_index][0]
    right_ax = axes[row_index][1]

    for ranker in RANKERS:
        color = RANKER_COLORS[ranker]
        left_vals = [metric_by_ranker[ranker][k][left_key] for k in ks]
        right_vals = [metric_by_ranker[ranker][k][right_key] for k in ks]
        left_ax.plot(ks, left_vals, marker="o", linewidth=2, color=color, label=ranker)
        right_ax.plot(
            ks, right_vals, marker="o", linewidth=2, color=color, label=ranker
        )

    left_ax.set_title(f"{row_title}: {left_title}")
    right_ax.set_title(f"{row_title}: {right_title}")
    left_ax.set_xlabel("k")
    right_ax.set_xlabel("k")
    left_ax.set_ylabel(left_ylabel)
    right_ax.set_ylabel(right_ylabel)
    left_ax.grid(alpha=0.3)
    right_ax.grid(alpha=0.3)
    left_ax.set_xticks(ks)
    right_ax.set_xticks(ks)


def save_metric_panels(
    output_path: Path,
    overall_metrics: dict[str, dict[int, dict[str, object]]],
    metrics_by_ara: dict[str, dict[str, dict[int, dict[str, object]]]],
    ks: list[int],
    left_key: str,
    left_title: str,
    left_ylabel: str,
    right_key: str,
    right_title: str,
    right_ylabel: str,
    figure_title: str,
) -> None:
    row_titles = ["Overall"] + sorted(metrics_by_ara)
    fig, axes = plt.subplots(
        nrows=len(row_titles),
        ncols=2,
        figsize=(14, 4 * len(row_titles)),
        constrained_layout=False,
    )
    if len(row_titles) == 1:
        axes = [axes]

    plot_line_panel(
        axes,
        0,
        "Overall",
        overall_metrics,
        ks,
        left_key,
        left_title,
        left_ylabel,
        right_key,
        right_title,
        right_ylabel,
    )
    for row_index, ara in enumerate(sorted(metrics_by_ara), start=1):
        plot_line_panel(
            axes,
            row_index,
            ara,
            metrics_by_ara[ara],
            ks,
            left_key,
            left_title,
            left_ylabel,
            right_key,
            right_title,
            right_ylabel,
        )

    handles, labels = axes[0][0].get_legend_handles_labels()
    fig.suptitle(figure_title, fontsize=14, y=0.985)
    fig.legend(
        handles,
        labels,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.955),
        ncol=2,
        frameon=False,
    )
    fig.tight_layout(rect=(0, 0, 1, 0.92))
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def save_pairwise_wins_plot(
    output_path: Path,
    pairwise_overall: dict[int, dict[str, dict[str, int]]],
    ks: list[int],
) -> None:
    fig, axes = plt.subplots(
        nrows=1, ncols=2, figsize=(14, 5), constrained_layout=False
    )
    plot_specs = [
        ("positive_hits", "Pairwise Wins: More TopAnswer Hits Is Better"),
        ("negative_hits", "Pairwise Wins: Fewer NeverShow Hits Is Better"),
    ]
    labels = [str(k) for k in ks]
    x = range(len(ks))
    width = 0.25

    for ax, (metric_key, title) in zip(axes, plot_specs):
        aragorn_vals = [pairwise_overall[k][metric_key]["ARAGORN_RANKER"] for k in ks]
        arax_vals = [pairwise_overall[k][metric_key]["ARAX_RANKER"] for k in ks]
        tie_zero_vals = [pairwise_overall[k][metric_key]["tie_zero"] for k in ks]
        tie_nonzero_vals = [pairwise_overall[k][metric_key]["tie_nonzero"] for k in ks]
        stack_totals = [
            max(a, b, tz + tn)
            for a, b, tz, tn in zip(
                aragorn_vals, arax_vals, tie_zero_vals, tie_nonzero_vals
            )
        ]
        ax.bar(
            [i - width for i in x],
            aragorn_vals,
            width=width,
            color=RANKER_COLORS["ARAGORN_RANKER"],
            label="ARAGORN_RANKER",
        )
        ax.bar(
            list(x),
            arax_vals,
            width=width,
            color=RANKER_COLORS["ARAX_RANKER"],
            label="ARAX_RANKER",
        )
        ax.bar(
            [i + width for i in x],
            tie_zero_vals,
            width=width,
            color="#c7c7c7",
            label="tie: both zero",
        )
        ax.bar(
            [i + width for i in x],
            tie_nonzero_vals,
            width=width,
            bottom=tie_zero_vals,
            color="#7f7f7f",
            label="tie: equal nonzero",
        )
        ax.set_title(title)
        ax.set_xlabel("k")
        ax.set_ylabel("(ARA, qid) groups won")
        ax.set_xticks(list(x), labels)
        ax.set_ylim(0, max(stack_totals) * 1.08 if stack_totals else 1)
        ax.grid(axis="y", alpha=0.3)

    handles, labels = axes[0].get_legend_handles_labels()
    fig.suptitle("Overall Pairwise Ranker Wins", fontsize=14, y=0.985)
    fig.text(
        0.5,
        0.94,
        "Light gray ties mean both rankers had zero hits; dark gray ties mean equal nonzero hits.",
        ha="center",
        va="center",
        fontsize=10,
    )
    fig.legend(
        handles,
        labels,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.915),
        ncol=4,
        frameon=False,
    )
    fig.tight_layout(rect=(0, 0, 1, 0.84))
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def save_pairwise_margin_plot(
    output_path: Path,
    pairwise_margin_overall: dict[
        int, dict[str, dict[str, dict[str, float | int | None]]]
    ],
    ks: list[int],
) -> None:
    fig, axes = plt.subplots(
        nrows=1, ncols=2, figsize=(14, 5), constrained_layout=False
    )
    plot_specs = [
        (
            "positive_hits",
            "Win Margin: Extra TopAnswer Hits",
            "Average extra TopAnswer hits among winning groups",
        ),
        (
            "negative_hits",
            "Win Margin: Fewer NeverShow Hits",
            "Average fewer NeverShow hits among winning groups",
        ),
    ]
    labels = [str(k) for k in ks]
    x = range(len(ks))
    width = 0.35

    for ax, (metric_key, title, ylabel) in zip(axes, plot_specs):
        aragorn_vals = [
            pairwise_margin_overall[k][metric_key]["ARAGORN_RANKER"]["mean_margin"] or 0
            for k in ks
        ]
        arax_vals = [
            pairwise_margin_overall[k][metric_key]["ARAX_RANKER"]["mean_margin"] or 0
            for k in ks
        ]
        ax.bar(
            [i - width / 2 for i in x],
            aragorn_vals,
            width=width,
            color=RANKER_COLORS["ARAGORN_RANKER"],
            label="ARAGORN_RANKER",
        )
        ax.bar(
            [i + width / 2 for i in x],
            arax_vals,
            width=width,
            color=RANKER_COLORS["ARAX_RANKER"],
            label="ARAX_RANKER",
        )
        for i, k in enumerate(ks):
            aragorn_wins = pairwise_margin_overall[k][metric_key]["ARAGORN_RANKER"][
                "win_count"
            ]
            arax_wins = pairwise_margin_overall[k][metric_key]["ARAX_RANKER"][
                "win_count"
            ]
            if aragorn_wins:
                ax.text(
                    i - width / 2,
                    aragorn_vals[i] + 0.03,
                    f"n={aragorn_wins}",
                    ha="center",
                    va="bottom",
                    fontsize=8,
                )
            if arax_wins:
                ax.text(
                    i + width / 2,
                    arax_vals[i] + 0.03,
                    f"n={arax_wins}",
                    ha="center",
                    va="bottom",
                    fontsize=8,
                )
        ymax = max(aragorn_vals + arax_vals) if (aragorn_vals or arax_vals) else 0
        ax.set_ylim(0, max(ymax * 1.2, 1))
        ax.set_title(title)
        ax.set_xlabel("k")
        ax.set_ylabel(ylabel)
        ax.set_xticks(list(x), labels)
        ax.grid(axis="y", alpha=0.3)

    handles, labels = axes[0].get_legend_handles_labels()
    fig.suptitle("Overall Pairwise Win Margins", fontsize=14, y=0.985)
    fig.text(
        0.5,
        0.94,
        "Bars show average margin among groups won; n labels show how many groups those averages come from.",
        ha="center",
        va="center",
        fontsize=10,
    )
    fig.legend(
        handles,
        labels,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.91),
        ncol=2,
        frameon=False,
    )
    fig.tight_layout(rect=(0, 0, 1, 0.84))
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def save_pairwise_margin_by_ara_plot(
    output_path: Path,
    pairwise_margin_by_ara: dict[
        str, dict[int, dict[str, dict[str, dict[str, float | int | None]]]]
    ],
    ks: list[int],
) -> None:
    aras = sorted(pairwise_margin_by_ara)
    if not aras:
        return

    fig, axes = plt.subplots(
        nrows=len(aras),
        ncols=2,
        figsize=(14, 4 * len(aras)),
        constrained_layout=False,
        squeeze=False,
    )
    labels = [str(k) for k in ks]
    x = range(len(ks))
    width = 0.35
    plot_specs = [
        (
            "positive_hits",
            "Extra TopAnswer Hits",
            "Average extra TopAnswer hits among winning groups",
        ),
        (
            "negative_hits",
            "Fewer NeverShow Hits",
            "Average fewer NeverShow hits among winning groups",
        ),
    ]

    for row_index, ara in enumerate(aras):
        for col_index, (metric_key, title_suffix, ylabel) in enumerate(plot_specs):
            ax = axes[row_index][col_index]
            aragorn_vals = [
                pairwise_margin_by_ara[ara][k][metric_key]["ARAGORN_RANKER"][
                    "mean_margin"
                ]
                or 0
                for k in ks
            ]
            arax_vals = [
                pairwise_margin_by_ara[ara][k][metric_key]["ARAX_RANKER"]["mean_margin"]
                or 0
                for k in ks
            ]
            ax.bar(
                [i - width / 2 for i in x],
                aragorn_vals,
                width=width,
                color=RANKER_COLORS["ARAGORN_RANKER"],
                label="ARAGORN_RANKER",
            )
            ax.bar(
                [i + width / 2 for i in x],
                arax_vals,
                width=width,
                color=RANKER_COLORS["ARAX_RANKER"],
                label="ARAX_RANKER",
            )
            for i, k in enumerate(ks):
                aragorn_wins = pairwise_margin_by_ara[ara][k][metric_key][
                    "ARAGORN_RANKER"
                ]["win_count"]
                arax_wins = pairwise_margin_by_ara[ara][k][metric_key]["ARAX_RANKER"][
                    "win_count"
                ]
                if aragorn_wins:
                    ax.text(
                        i - width / 2,
                        aragorn_vals[i] + 0.03,
                        f"n={aragorn_wins}",
                        ha="center",
                        va="bottom",
                        fontsize=8,
                    )
                if arax_wins:
                    ax.text(
                        i + width / 2,
                        arax_vals[i] + 0.03,
                        f"n={arax_wins}",
                        ha="center",
                        va="bottom",
                        fontsize=8,
                    )
            ymax = max(aragorn_vals + arax_vals) if (aragorn_vals or arax_vals) else 0
            ax.set_ylim(0, max(ymax * 1.2, 1))
            ax.set_title(f"{ara}: {title_suffix}")
            ax.set_xlabel("k")
            ax.set_ylabel(ylabel)
            ax.set_xticks(list(x), labels)
            ax.grid(axis="y", alpha=0.3)

    handles, legend_labels = axes[0][0].get_legend_handles_labels()
    fig.suptitle("Pairwise Win Margins by ARA", fontsize=14, y=0.985)
    fig.text(
        0.5,
        0.95,
        "Bar height = average win margin; n labels = number of winning groups contributing to that average.",
        ha="center",
        va="center",
        fontsize=10,
    )
    fig.legend(
        handles,
        legend_labels,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.925),
        ncol=2,
        frameon=False,
    )
    fig.tight_layout(rect=(0, 0, 1, 0.90))
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def generate_plots(
    plot_dir: Path,
    overall_metrics: dict[str, dict[int, dict[str, object]]],
    pairwise_overall: dict[int, dict[str, dict[str, int]]],
    pairwise_margin_overall: dict[
        int, dict[str, dict[str, dict[str, float | int | None]]]
    ],
    pairwise_margin_by_ara: dict[
        str, dict[int, dict[str, dict[str, dict[str, float | int | None]]]]
    ],
    metrics_by_ara: dict[str, dict[str, dict[int, dict[str, object]]]],
    ks: list[int],
) -> list[str]:
    plot_dir.mkdir(parents=True, exist_ok=True)
    output_files = []

    topanswer_path = plot_dir / "topanswer_comparison.png"
    save_metric_panels(
        topanswer_path,
        overall_metrics,
        metrics_by_ara,
        ks,
        left_key="positive_hits",
        left_title="TopAnswer hits@k",
        left_ylabel="TopAnswer hits in top-k",
        right_key="positive_recall_micro",
        right_title="TopAnswer recall@k",
        right_ylabel="Micro recall",
        figure_title="TopAnswer Comparison Across Rankers",
    )
    output_files.append(str(topanswer_path))

    nevershow_path = plot_dir / "nevershow_comparison.png"
    save_metric_panels(
        nevershow_path,
        overall_metrics,
        metrics_by_ara,
        ks,
        left_key="negative_hits",
        left_title="NeverShow hits@k",
        left_ylabel="NeverShow hits in top-k",
        right_key="negative_specificity_micro",
        right_title="NeverShow specificity@k",
        right_ylabel="Micro specificity",
        figure_title="NeverShow Comparison Across Rankers",
    )
    output_files.append(str(nevershow_path))

    f1_path = plot_dir / "f1_comparison.png"
    save_metric_panels(
        f1_path,
        overall_metrics,
        metrics_by_ara,
        ks,
        left_key="precision_micro",
        left_title="Precision@k",
        left_ylabel="Micro precision",
        right_key="f1_micro",
        right_title="F1@k",
        right_ylabel="Micro F1",
        figure_title="Precision and F1 Comparison Across Rankers",
    )
    output_files.append(str(f1_path))

    pairwise_path = plot_dir / "pairwise_wins.png"
    save_pairwise_wins_plot(pairwise_path, pairwise_overall, ks)
    output_files.append(str(pairwise_path))

    pairwise_margin_path = plot_dir / "pairwise_win_margins.png"
    save_pairwise_margin_plot(pairwise_margin_path, pairwise_margin_overall, ks)
    output_files.append(str(pairwise_margin_path))

    pairwise_margin_by_ara_path = plot_dir / "pairwise_win_margins_by_ara.png"
    save_pairwise_margin_by_ara_plot(
        pairwise_margin_by_ara_path, pairwise_margin_by_ara, ks
    )
    output_files.append(str(pairwise_margin_by_ara_path))

    return output_files


def build_json_payload(
    args: argparse.Namespace,
    metadata: dict[str, object],
    duplicate_report_file: str,
    duplicate_report: list[dict[str, object]],
    overall_metrics: dict[str, dict[int, dict[str, object]]],
    pairwise_overall: dict[int, dict[str, dict[str, int]]],
    pairwise_margin_overall: dict[
        int, dict[str, dict[str, dict[str, float | int | None]]]
    ],
    metrics_by_ara: dict[str, dict[str, dict[int, dict[str, object]]]],
    pairwise_by_ara: dict[str, dict[int, dict[str, dict[str, int]]]],
    pairwise_margin_by_ara: dict[
        str, dict[int, dict[str, dict[str, dict[str, float | int | None]]]]
    ],
    strongest_positive_examples: dict[str, list[dict[str, object]]],
    strongest_negative_examples: dict[str, dict[str, list[dict[str, object]]]],
    plot_files: list[str],
) -> dict[str, object]:
    return {
        "workbook": str(args.workbook),
        "sheet": metadata["sheet"],
        "ks": args.ks,
        "positive_labels": sorted(args.positive_labels),
        "negative_labels": sorted(args.negative_labels),
        "dataset_summary": metadata,
        "duplicate_report_file": duplicate_report_file,
        "duplicate_query_edge_ids": duplicate_report,
        "overall_metrics": overall_metrics,
        "pairwise_overall": pairwise_overall,
        "pairwise_margin_overall": pairwise_margin_overall,
        "metrics_by_ara": metrics_by_ara,
        "pairwise_by_ara": pairwise_by_ara,
        "pairwise_margin_by_ara": pairwise_margin_by_ara,
        "strongest_positive_examples": strongest_positive_examples,
        "strongest_negative_examples": strongest_negative_examples,
        "plot_files": plot_files,
    }


def main() -> None:
    args = parse_args()
    ks = sorted(set(args.ks))
    positive_labels = set(args.positive_labels)
    negative_labels = set(args.negative_labels)

    if args.plot_dir is None:
        plot_dir = args.workbook.parent / f"{args.workbook.stem}_plots"
    else:
        plot_dir = args.plot_dir

    if args.duplicate_report is None:
        duplicate_report_path = (
            args.workbook.parent / f"{args.workbook.stem}_duplicate_edge_ids.json"
        )
    else:
        duplicate_report_path = args.duplicate_report

    records, metadata, duplicate_report = load_edge_records(args.workbook, args.sheet)
    (
        overall_metrics,
        pairwise_overall,
        pairwise_margin_overall,
        metrics_by_ara,
        pairwise_by_ara,
        pairwise_margin_by_ara,
        strongest_positive_examples,
        strongest_negative_examples,
    ) = aggregate_metric_table(records, ks, positive_labels, negative_labels)
    plot_files = generate_plots(
        plot_dir,
        overall_metrics,
        pairwise_overall,
        pairwise_margin_overall,
        pairwise_margin_by_ara,
        metrics_by_ara,
        ks,
    )
    duplicate_report_path.write_text(
        json.dumps(duplicate_report, indent=2), encoding="utf-8"
    )

    print_dataset_summary(metadata, records)
    print()
    print_duplicate_summary(duplicate_report)
    print_metric_block("Overall metrics", overall_metrics, ks)
    print_pairwise_block("Overall pairwise wins", pairwise_overall, ks)
    print_pairwise_margin_block(
        "Overall pairwise win margins", pairwise_margin_overall, ks
    )
    for ara in sorted(metrics_by_ara):
        print_metric_block(f"Metrics for {ara}", metrics_by_ara[ara], ks)
        print_pairwise_block(f"Pairwise wins for {ara}", pairwise_by_ara[ara], ks)
    print_examples(
        "Largest TopAnswer hit differences",
        strongest_positive_examples,
        args.show_group_examples,
        "positive_hit_delta",
    )
    print_examples(
        "Largest NeverShow differences where fewer is better",
        strongest_negative_examples["fewest_nevershow"],
        args.show_group_examples,
        "negative_hit_delta",
    )
    print_examples(
        "Largest NeverShow differences where more is worse",
        strongest_negative_examples["most_nevershow"],
        args.show_group_examples,
        "negative_hit_delta",
    )
    print("Generated plots")
    for plot_file in plot_files:
        print(f"  {plot_file}")
    print(f"Duplicate report\n  {duplicate_report_path}")

    if args.json_output:
        payload = build_json_payload(
            args,
            metadata,
            str(duplicate_report_path),
            duplicate_report,
            overall_metrics,
            pairwise_overall,
            pairwise_margin_overall,
            metrics_by_ara,
            pairwise_by_ara,
            pairwise_margin_by_ara,
            strongest_positive_examples,
            strongest_negative_examples,
            plot_files,
        )
        args.json_output.write_text(json.dumps(payload, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
